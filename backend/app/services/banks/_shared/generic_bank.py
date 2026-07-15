"""
Generic bank processing helpers for the new bank modules.

This module provides a reusable parser / validator / classifier / report pipeline
so new bank folders can stay small and still follow the same package structure as
HDFC, Axis, ICICI, Kotak, and SBI.
"""

from __future__ import annotations

import json
import logging
import os
import re
import time
import uuid
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from .category_registry import normalize_category

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

try:
    from pypdf import PdfReader  # type: ignore
except Exception:  # pragma: no cover - optional runtime dependency fallback
    try:
        from PyPDF2 import PdfReader  # type: ignore
    except Exception:  # pragma: no cover - optional runtime dependency fallback
        PdfReader = None  # type: ignore

logger = logging.getLogger(__name__)


@dataclass
class GenericBankConfig:
    bank_key: str
    bank_name: str
    file_prefix: str
    markers: List[str] = field(default_factory=list)
    support_aliases: List[str] = field(default_factory=list)


@dataclass
class GenericTransaction:
    date: str
    description: str
    debit: float = 0.0
    credit: float = 0.0
    balance: Optional[float] = None
    reference: Optional[str] = None
    category: str = "Others"
    confidence: float = 0.0
    source: str = "parser"
    matched_rule: Optional[str] = None
    matched_keyword: Optional[str] = None
    raw_line: Optional[str] = None
    is_recurring: bool = False

    def to_dict(self) -> Dict[str, Any]:
        return {
            "date": self.date,
            "description": self.description,
            "debit": self.debit,
            "credit": self.credit,
            "balance": self.balance,
            "reference": self.reference,
            "category": self.category,
            "confidence": self.confidence,
            "source": self.source,
            "matched_rule": self.matched_rule,
            "matched_keyword": self.matched_keyword,
            "raw_line": self.raw_line,
            "is_recurring": self.is_recurring,
        }


@dataclass
class GenericParseResult:
    status: str
    transactions: List[GenericTransaction]
    text_content: str = ""
    metadata: Dict[str, Any] = field(default_factory=dict)
    error_code: Optional[str] = None
    error_message: Optional[str] = None

    def to_dict(self) -> Dict[str, Any]:
        return {
            "status": self.status,
            "transactions": [txn.to_dict() for txn in self.transactions],
            "metadata": self.metadata,
            "error_code": self.error_code,
            "error_message": self.error_message,
        }


@dataclass
class GenericStructureMetadata:
    account_number: Optional[str] = None
    account_holder: Optional[str] = None
    statement_from: Optional[str] = None
    statement_to: Optional[str] = None
    opening_balance: Optional[float] = None
    closing_balance: Optional[float] = None
    ifsc: Optional[str] = None
    # Rich statement-summary / KYC fields (HDFC-parity)
    dr_count: Optional[int] = None
    cr_count: Optional[int] = None
    total_debits: Optional[float] = None
    total_credits: Optional[float] = None
    branch: Optional[str] = None
    account_type: Optional[str] = None
    email: Optional[str] = None
    mobile: Optional[str] = None
    address: Optional[str] = None
    pan: Optional[str] = None
    account_open_date: Optional[str] = None
    joint_holders: Optional[str] = None
    customer_id: Optional[str] = None
    micr: Optional[str] = None
    crn: Optional[str] = None

    def to_dict(self) -> Dict[str, Any]:
        from .rich_metadata import metadata_to_rich_dict
        return metadata_to_rich_dict(self)

    @property
    def expected_transaction_count(self) -> Optional[int]:
        if self.dr_count is not None and self.cr_count is not None:
            return int(self.dr_count) + int(self.cr_count)
        return None



@dataclass
class GenericStructureResult:
    is_valid: bool
    confidence: float
    metadata: GenericStructureMetadata
    text_content: str
    error_code: Optional[str] = None
    error_message: Optional[str] = None

    def to_dict(self) -> Dict[str, Any]:
        return {
            "is_valid": self.is_valid,
            "confidence": self.confidence,
            "metadata": self.metadata.to_dict(),
            "error_code": self.error_code,
            "error_message": self.error_message,
        }


@dataclass
class GenericProcessingMetrics:
    total_time_ms: float = 0.0
    step_timings: Dict[str, float] = field(default_factory=dict)
    transaction_count: int = 0
    classified_count: int = 0
    unclassified_count: int = 0
    recurring_count: int = 0
    reconciliation_passed: bool = False


@dataclass
class GenericProcessingResult:
    status: str
    excel_path: Optional[str]
    transactions: List[Dict[str, Any]]
    aggregation: Any
    metrics: GenericProcessingMetrics
    bank_key: str = ""
    error_message: Optional[str] = None
    error_code: Optional[str] = None

    def to_dict(self) -> Dict[str, Any]:
        stats = {
            "total_transactions": self.metrics.transaction_count,
            "classified": self.metrics.classified_count,
            "others": self.metrics.unclassified_count,
            "recurring": self.metrics.recurring_count,
            "coverage_percent": round(
                self.metrics.classified_count / max(self.metrics.transaction_count, 1) * 100, 1
            ),
        }
        return {
            "status": self.status,
            "bank_key": self.bank_key,
            "excel_path": self.excel_path,
            "stats": stats,
            "validation": {
                "reconciliation_passed": self.metrics.reconciliation_passed,
            },
            "performance": self.metrics.step_timings,
            "error": {
                "message": self.error_message,
                "code": self.error_code,
            } if self.error_message else None,
        }


class GenericProcessorError(Exception):
    def __init__(self, message: str, stage: str, error_code: str, details: dict = None):
        self.stage = stage
        self.error_code = error_code
        self.details = details or {}
        super().__init__(f"[{stage}] {message}")


class GenericStructureError(Exception):
    def __init__(self, message: str, error_code: str, details: dict = None):
        self.error_code = error_code
        self.details = details or {}
        super().__init__(message)


class GenericParseError(Exception):
    def __init__(self, message: str, error_code: str = "UNKNOWN", details: dict = None):
        self.error_code = error_code
        self.details = details or {}
        super().__init__(message)


class GenericValidationError(Exception):
    def __init__(self, message: str, error_code: str = "UNKNOWN", details: dict = None):
        self.error_code = error_code
        self.details = details or {}
        super().__init__(message)


class GenericReconciliationError(Exception):
    def __init__(self, message: str, error_code: str = "UNKNOWN", details: dict = None):
        self.error_code = error_code
        self.details = details or {}
        super().__init__(message)


class GenericClassifier:
    """Generic rule-based classifier backed by the shared words.json structure."""

    def __init__(self, bank_config: GenericBankConfig, keywords_file: Optional[str] = None):
        self.bank_config = bank_config
        db = self._load_db(bank_config.bank_key, keywords_file)
        self._meta = db.get("metadata", {})
        self._build_normalization(db.get("text_normalization", {}))
        self._build_entity_lookup(db.get("entity_interpretation", {}))
        self._build_pattern_sets(db.get("pattern_detection", {}))
        logger.info(
            "%s classifier loaded: %d entity aliases — v%s",
            bank_config.bank_name,
            len(self._entity_lookup),
            self._meta.get("version", "?"),
        )

    @staticmethod
    def _repo_root() -> Path:
        return Path(__file__).resolve().parents[5]

    def _load_db(self, bank_key: str, path: Optional[str]) -> Dict[str, Any]:
        alternate_folders = {
            "bank_of_baroda": ["bank of baroda"],
            "karnataka": ["karnataka bank"],
        }
        candidates = [
            path,
            f"/app/samples/{bank_key}/output/words.json",
            "/app/keywords.json",
            "/app/words.json",
            str(self._repo_root() / "backend" / "words.json"),
            str(self._repo_root() / "words.json"),
            str(self._repo_root() / "samples" / bank_key / "output" / "words.json"),
        ]
        for folder in alternate_folders.get(bank_key, []):
            candidates.append(str(self._repo_root() / "samples" / folder / "output" / "words.json"))
        for candidate in candidates:
            if not candidate:
                continue
            try:
                with open(candidate, "r", encoding="utf-8") as fh:
                    return json.load(fh)
            except Exception:
                continue
        logger.warning("No keyword database found for %s; using empty rules", bank_key)
        return {}

    def _build_normalization(self, tnorm: Dict[str, Any]) -> None:
        self._strip_regex = [
            re.compile(p, re.IGNORECASE)
            for p in tnorm.get("strip_patterns", [])
        ]
        self._replace_rules: List[Tuple[re.Pattern, str]] = []
        for old, new in tnorm.get("replace_rules", {}).items():
            if old.isalpha() and len(old) <= 4:
                pat = re.compile(r"\b" + re.escape(old) + r"\b", re.IGNORECASE)
            else:
                pat = re.compile(re.escape(old), re.IGNORECASE)
            self._replace_rules.append((pat, new))

    def _build_entity_lookup(self, entity_interp: Dict[str, Any]) -> None:
        self._entity_lookup: Dict[str, Dict[str, Any]] = {}
        for group_name, group_data in entity_interp.items():
            if group_name.startswith("_") or not isinstance(group_data, dict):
                continue
            for alias, cfg in group_data.items():
                if alias.startswith("_") or not isinstance(cfg, dict):
                    continue
                key = alias.lower()
                if "credit" in cfg and "debit" in cfg:
                    credit_cfg = cfg["credit"]
                    debit_cfg = cfg["debit"]
                    credit_cat = credit_cfg.get("category", "TRANSFER_IN")
                    debit_cat = debit_cfg.get("category", "TRANSFER_OUT")
                    priority = max(credit_cfg.get("priority", 50), debit_cfg.get("priority", 50))
                    group = credit_cfg.get("group", group_name)
                elif "category" in cfg:
                    credit_cat = cfg["category"]
                    debit_cat = cfg["category"]
                    priority = cfg.get("priority", 50)
                    group = cfg.get("group", group_name)
                else:
                    continue
                existing = self._entity_lookup.get(key)
                if existing is None or priority > existing["priority"]:
                    self._entity_lookup[key] = {
                        "credit_cat": credit_cat,
                        "debit_cat": debit_cat,
                        "priority": priority,
                        "group": group,
                    }
        self._sorted_aliases: List[str] = sorted(self._entity_lookup.keys(), key=len, reverse=True)

    def _build_pattern_sets(self, pd_cfg: Dict[str, Any]) -> None:
        def ll(key: str) -> List[str]:
            return [p.lower() for p in pd_cfg.get(key, [])]

        self._upi_handles: List[str] = ll("upi_handle")
        self._loan_patterns: List[str] = ll("loan_patterns")
        self._refund_patterns: List[str] = ll("refund_patterns")
        self._bill_patterns: List[str] = ll("bill_payment_patterns")
        self._insurance_patterns: List[str] = ll("insurance_patterns")
        self._transfer_patterns: List[str] = ll("transfer_patterns")
        self._settlement_patterns: List[str] = ll("settlement_patterns")
        self._fuel_patterns: List[str] = ll("fuel_travel_patterns")
        self._interest_tokens: List[str] = [
            "interest credit", "interest debit", "int credit", "int debit",
            "int cr", "int dr", "interest earned", "interest charged",
        ]
        self._atm_tokens: List[str] = [
            "atm", "atm withdrawal", "cash withdrawal", "cash wdl", "wdl atm",
        ]
        self._gst_tokens: List[str] = ["cgst", "sgst", "igst", " gst"]
        self._charge_tokens: List[str] = [
            "bank charges", "service charge", "processing fee", "annual fee",
            "card fee", "late payment charge", "sms charge", "minimum balance",
        ]
        self._refund_tokens: List[str] = list({*self._refund_patterns, "refund", "reversal", "chargeback", "cashback", "cash back"})
        self._salary_tokens: List[str] = ["salary", "payroll", "wages", "sal credit", "sal cr", "monthly salary"]
        self._ach_tokens: List[str] = ["ach d", "ach debit", "nach", "ecs ", "auto debit", "standing instruction"]

    def _normalize(self, text: str) -> str:
        if not text:
            return ""
        s = text.lower().strip()
        for regex in self._strip_regex:
            s = regex.sub(" ", s)
        for regex, replacement in self._replace_rules:
            s = regex.sub(replacement, s)
        s = re.sub(r"\s+", " ", s)
        return s.strip()

    def classify(self, transactions: List[Dict[str, Any]]) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
        classified: List[Dict[str, Any]] = []
        unclassified: List[Dict[str, Any]] = []

        for txn in transactions:
            # Skip invalid transactions that aren't dictionaries
            if not isinstance(txn, dict):
                continue
            result = self._classify_single(txn)
            txn_copy = dict(txn)
            txn_copy["category"] = normalize_category(
                result[0],
                is_debit=bool(txn.get("debit")),
            )
            txn_copy["confidence"] = result[1]
            txn_copy["source"] = result[2]
            txn_copy["matched_rule"] = result[3]
            txn_copy["matched_keyword"] = result[4]
            if result[0].startswith("Others"):
                unclassified.append(txn_copy)
            else:
                classified.append(txn_copy)

        return classified, unclassified

    def _classify_single(self, txn: Dict[str, Any]) -> Tuple[str, float, str, Optional[str], Optional[str]]:
        description = self._normalize(str(txn.get("description") or txn.get("narration") or ""))
        is_debit = float(txn.get("debit") or 0) > 0
        rules = self._debit_rules() if is_debit else self._credit_rules()
        default_category = "Others Debit" if is_debit else "Others Credit"

        for category, compiled in rules.items():
            for keyword in compiled["exact"]:
                if keyword in description:
                    return category, 0.99, "rule_engine", "exact_keyword", keyword

        for category, compiled in rules.items():
            for pattern in compiled["patterns"]:
                if pattern.search(description):
                    return category, 0.95, "rule_engine", "pattern_match", pattern.pattern

        if is_debit and "upi" in description:
            for merchant, category in self._upi_merchant_map().items():
                if merchant in description:
                    return category, 0.85, "rule_engine", "upi_merchant", merchant

        if any(token in description for token in self._salary_tokens):
            return ("Salary", 0.95, "rule_engine", "salary_token", "salary")
        if any(token in description for token in self._refund_tokens):
            return ("Refund", 0.90, "rule_engine", "refund_token", "refund")
        if any(token in description for token in self._charge_tokens):
            return ("Bank Charges", 0.88, "rule_engine", "charges_token", "charges")

        # ── Entity lookup from words.json ─────────────────────────────────────
        # Walk longest-first aliases so "bajaj finance" beats "bajaj"
        for alias in self._sorted_aliases:
            if alias in description:
                entry = self._entity_lookup[alias]
                raw_cat = entry["debit_cat"] if is_debit else entry["credit_cat"]
                return raw_cat, 0.82, "entity_lookup", "words_json", alias

        # ── Pattern sets from words.json ──────────────────────────────────────
        if any(p in description for p in self._loan_patterns):
            return ("Loan Payment" if is_debit else "Loan Disbursed", 0.88, "pattern_set", "loan_pattern", None)
        if any(p in description for p in self._bill_patterns):
            return ("Bill Payment", 0.85, "pattern_set", "bill_pattern", None)
        if any(p in description for p in self._insurance_patterns):
            return ("Insurance", 0.85, "pattern_set", "insurance_pattern", None)
        if any(p in description for p in self._fuel_patterns):
            return ("Fuel", 0.85, "pattern_set", "fuel_pattern", None)
        if any(p in description for p in self._transfer_patterns):
            return ("Transfer", 0.80, "pattern_set", "transfer_pattern", None)
        if any(p in description for p in self._settlement_patterns):
            return ("Transfer", 0.80, "pattern_set", "settlement_pattern", None)
        if any(h in description for h in self._upi_handles):
            return ("Transfer" if is_debit else "Others Credit", 0.78, "pattern_set", "upi_handle", None)

        return default_category, 0.5, "rule_engine", "default", None

    def get_category_stats(self) -> Dict[str, Any]:
        """Return lightweight classifier metadata for logging and diagnostics."""
        return {
            "bank_key": self.bank_config.bank_key,
            "bank_name": self.bank_config.bank_name,
            "entity_aliases": len(getattr(self, "_entity_lookup", {})),
            "upi_handles": len(getattr(self, "_upi_handles", [])),
            "loan_patterns": len(getattr(self, "_loan_patterns", [])),
            "refund_patterns": len(getattr(self, "_refund_patterns", [])),
        }

    def get_all_categories(self) -> Dict[str, List[str]]:
        """Return all possible display categories grouped by direction."""
        debit_rules = self._debit_rules()
        credit_rules = self._credit_rules()
        return {
            "credit": list(credit_rules.keys()),
            "debit": list(debit_rules.keys()),
        }

    def _debit_rules(self) -> Dict[str, Dict[str, Any]]:
        return {
            "ATM Withdrawal": {"exact": {"atm", "cash withdrawal", "wdl atm"}, "patterns": [re.compile(r"atm.*withdraw", re.IGNORECASE)]},
            "Loan EMI": {"exact": {"loan emi", "loan payment", "emi payment", "emi debit"}, "patterns": [re.compile(r"\bemi\b", re.IGNORECASE), re.compile(r"loan.*emi", re.IGNORECASE), re.compile(r".*loan.*", re.IGNORECASE)]},
            "Fuel": {"exact": {"petrol", "diesel", "fuel"}, "patterns": [re.compile(r".*fuel.*", re.IGNORECASE)]},
            "Shopping": {"exact": {"amazon", "flipkart", "myntra", "ajio"}, "patterns": [re.compile(r".*shopping.*", re.IGNORECASE)]},
            "Food & Dining": {"exact": {"swiggy", "zomato", "restaurant", "cafe"}, "patterns": [re.compile(r".*food.*", re.IGNORECASE)]},
            "Utilities": {"exact": {"electricity", "water", "gas", "mobile recharge"}, "patterns": [re.compile(r".*bill.*", re.IGNORECASE)]},
            "Transfer": {"exact": {"neft", "rtgs", "imps", "upi"}, "patterns": [re.compile(r".*transfer.*", re.IGNORECASE)]},
        }

    def _credit_rules(self) -> Dict[str, Dict[str, Any]]:
        return {
            "Salary": {"exact": {"salary", "payroll", "wages"}, "patterns": [re.compile(r".*salary.*", re.IGNORECASE)]},
            "Interest": {"exact": {"interest"}, "patterns": [re.compile(r".*interest.*", re.IGNORECASE)]},
            "Refund": {"exact": {"refund", "reversal", "cashback"}, "patterns": [re.compile(r".*refund.*", re.IGNORECASE)]},
            "Transfer In": {"exact": {"neft", "rtgs", "imps", "upi"}, "patterns": [re.compile(r".*credit.*", re.IGNORECASE)]},
        }

    def _upi_merchant_map(self) -> Dict[str, str]:
        return {
            "swiggy": "Food & Dining",
            "zomato": "Food & Dining",
            "amazon": "Shopping",
            "flipkart": "Shopping",
            "uber": "Transport",
            "ola": "Transport",
            "paytm": "Others Debit",
            "phonepe": "Others Debit",
            "gpay": "Others Debit",
        }

    def get_statistics(self) -> Dict[str, Any]:
        return {
            "entity_aliases": len(self._entity_lookup),
            "version": self._meta.get("version", "?"),
        }


class GenericStructureValidator:
    def __init__(self, bank_config: GenericBankConfig):
        self.bank_config = bank_config
        self.logger = logging.getLogger(f"{__name__}.{self.__class__.__name__}")
        self._markers = [m.lower() for m in bank_config.markers]

        self._account_patterns = [
            re.compile(r"(?:account|a/c|acct)\s*(?:no|number)?\s*[:\-]?\s*([\dXx\-]{8,22})", re.IGNORECASE),
            re.compile(r"(?:savings|current)\s*account\s*[:\-]?\s*([\dXx\-]{8,22})", re.IGNORECASE),
        ]
        self._period_patterns = [
            re.compile(r"(?:from|period from)\s*[:\-]?\s*(\d{2}[\-/]\d{2}[\-/]\d{2,4})\s*(?:to|-)\s*(\d{2}[\-/]\d{2}[\-/]\d{2,4})", re.IGNORECASE),
        ]
        self._ifsc_patterns = [re.compile(r"\b([A-Z]{4}0[A-Z0-9]{6})\b")]
        self._opening_patterns = [re.compile(r"opening\s*balance\s*[:\-]?\s*([\d,]+(?:\.\d{1,2})?)", re.IGNORECASE)]
        self._closing_patterns = [re.compile(r"closing\s*balance\s*[:\-]?\s*([\d,]+(?:\.\d{1,2})?)", re.IGNORECASE)]

    def validate(self, text_content: str, first_page_text: str = "") -> GenericStructureResult:
        raw_header = first_page_text or text_content[:7000] or ""
        header_text = raw_header.lower()
        body_text = text_content.lower()
        confidence = self._check_markers(header_text)
        if self.bank_config.bank_key != "unknown" and confidence < 0.25:
            raise GenericStructureError(
                f"PDF does not appear to be a {self.bank_config.bank_name} statement",
                error_code=f"NOT_{self.bank_config.bank_key.upper()}_STATEMENT",
                details={"confidence": confidence},
            )

        # Extract with lowercased text for basic patterns, then enrich from original case
        metadata = self._extract_metadata(body_text, header_text)
        try:
            from .rich_metadata import enrich_statement_metadata
            enrich_statement_metadata(metadata, text_content, raw_header)
        except Exception:
            pass
        if not self._has_transaction_structure(body_text):
            if self.bank_config.bank_key != "unknown":
                raise GenericStructureError(
                    "Could not identify a transaction table in this statement",
                    error_code="NO_TRANSACTION_TABLE",
                    details={},
                )

        return GenericStructureResult(
            is_valid=True,
            confidence=confidence,
            metadata=metadata,
            text_content=text_content,
        )


    def _check_markers(self, text: str) -> float:
        # GenericStructureValidator lowercases inputs; keep raw for enricher via originals
        # Callers pass lowercased text here — re-enrich from original if available is done by banks.
        if not self._markers:
            return 0.3 if text else 0.0
        found = sum(1 for marker in self._markers if marker in text)
        return min(found / max(len(self._markers), 2), 1.0)

    def _has_transaction_structure(self, text: str) -> bool:
        date_hits = len(re.findall(r"\b\d{2}[\-/]\d{2}[\-/]\d{2,4}\b", text))
        header_hits = any(
            kw in text for kw in (
                "date", "particular", "description", "narration", "debit", "credit", "balance"
            )
        )
        return date_hits > 2 or header_hits

    def _extract_metadata(self, full_text: str, header_text: str) -> GenericStructureMetadata:
        metadata = GenericStructureMetadata()
        for pattern in self._account_patterns:
            match = pattern.search(full_text) or pattern.search(header_text)
            if match:
                metadata.account_number = match.group(1)
                break
        for pattern in self._period_patterns:
            match = pattern.search(full_text)
            if match:
                metadata.statement_from = match.group(1)
                metadata.statement_to = match.group(2)
                break
        for pattern in self._ifsc_patterns:
            match = pattern.search(full_text)
            if match:
                metadata.ifsc = match.group(1)
                break
        for pattern in self._opening_patterns:
            match = pattern.search(full_text)
            if match:
                metadata.opening_balance = self._parse_amount(match.group(1))
                break
        for pattern in self._closing_patterns:
            match = pattern.search(full_text)
            if match:
                metadata.closing_balance = self._parse_amount(match.group(1))
                break
        try:
            from .rich_metadata import enrich_statement_metadata
            enrich_statement_metadata(metadata, full_text, header_text)
        except Exception:
            pass
        return metadata


    @staticmethod
    def _parse_amount(amount_str: str) -> Optional[float]:
        try:
            return float(amount_str.replace(",", ""))
        except Exception:
            return None


class GenericParser:
    def __init__(self, bank_config: GenericBankConfig):
        self.bank_config = bank_config
        self.logger = logging.getLogger(f"{__name__}.{self.__class__.__name__}")
        self.structure_validator = GenericStructureValidator(bank_config)

    def parse(self, file_path: str, text_content: str = "") -> GenericParseResult:
        try:
            text = text_content or self._extract_text(file_path)
            if not text.strip():
                raise GenericParseError("Could not extract any text from PDF", error_code="EMPTY_TEXT")

            structure = self.structure_validator.validate(text)
            transactions = self._parse_transactions(text)
            return GenericParseResult(
                status="success",
                transactions=transactions,
                text_content=text,
                metadata=structure.metadata.to_dict(),
            )
        except GenericParseError:
            raise
        except GenericStructureError as exc:
            raise GenericParseError(str(exc), error_code=getattr(exc, "error_code", "STRUCTURE_ERROR"), details=getattr(exc, "details", {}))
        except Exception as exc:
            raise GenericParseError(str(exc), error_code="PARSING_ERROR", details={})

    def _extract_text(self, file_path: str) -> str:
        if PdfReader is not None:
            try:
                reader = PdfReader(file_path)
                pages = []
                for page in reader.pages:
                    pages.append(page.extract_text() or "")
                text = "\n".join(pages)
                if text.strip():
                    return text
            except Exception:
                pass

        try:
            import fitz  # type: ignore

            doc = fitz.open(file_path)
            pages = [page.get_text("text") or "" for page in doc]
            doc.close()
            text = "\n".join(pages)
            if text.strip():
                return text
        except Exception:
            pass

        try:
            import pdfplumber  # type: ignore

            pages = []
            with pdfplumber.open(file_path) as pdf:
                for page in pdf.pages:
                    pages.append(page.extract_text() or "")
            return "\n".join(pages)
        except Exception:
            return ""

    def _parse_transactions(self, text: str) -> List[GenericTransaction]:
        lines = [line.strip() for line in text.splitlines() if line.strip()]
        transactions: List[GenericTransaction] = []
        current: List[str] = []

        for line in lines:
            if self._looks_like_transaction_start(line):
                if current:
                    txn = self._build_transaction(current)
                    if txn:
                        transactions.append(txn)
                current = [line]
            elif current and not self._should_skip_line(line):
                current.append(line)

        if current:
            txn = self._build_transaction(current)
            if txn:
                transactions.append(txn)

        if not transactions:
            # Fallback: try to turn any date-bearing line into a transaction.
            for line in lines:
                if self._looks_like_transaction_start(line):
                    txn = self._build_transaction([line])
                    if txn:
                        transactions.append(txn)

        return transactions

    def _looks_like_transaction_start(self, line: str) -> bool:
        return bool(re.match(r"^(?:\d{2}[\-/]\d{2}[\-/]\d{2,4}|\d{1,2}\s+[A-Za-z]{3}\s+\d{2,4}|[A-Za-z]{3}\s+\d{1,2},?\s+\d{4})\b", line))

    def _should_skip_line(self, line: str) -> bool:
        lowered = line.lower()
        return any(
            token in lowered
            for token in (
                "page ", "statement summary", "opening balance", "closing balance",
                "account summary", "generated on", "branch code", "customer id",
            )
        )

    def _build_transaction(self, lines: List[str]) -> Optional[GenericTransaction]:
        text = " ".join(lines)
        date_match = re.match(r"^(\d{2}[\-/]\d{2}[\-/]\d{2,4}|\d{1,2}\s+[A-Za-z]{3}\s+\d{2,4}|[A-Za-z]{3}\s+\d{1,2},?\s+\d{4})\b", text)
        if not date_match:
            return None
        date_text = self._normalize_date(date_match.group(1))
        remainder = text[date_match.end():].strip()
        numbers = self._extract_amounts(remainder)
        description = self._clean_description(remainder)
        debit = credit = 0.0
        balance = numbers[-1] if numbers else None
        txn_amount = numbers[-2] if len(numbers) >= 2 else (numbers[0] if numbers else None)

        desc_lower = description.lower()
        if txn_amount is not None:
            if any(token in desc_lower for token in ("salary", "credit", "refund", "cashback", "interest", "deposit")):
                credit = abs(txn_amount)
            elif any(token in desc_lower for token in ("debit", "withdraw", "purchase", "payment", "transfer", "upi", "imps", "neft", "rtgs", "emi", "charge")):
                debit = abs(txn_amount)
            elif txn_amount < 0:
                debit = abs(txn_amount)
            else:
                debit = abs(txn_amount) if len(numbers) <= 2 else 0.0
                if debit == 0.0:
                    credit = abs(txn_amount)

        if not description:
            description = remainder

        if not description and not numbers:
            return None

        reference = None
        ref_match = re.search(r"\b(?:ref|utr|rrn|txn|transaction)\s*[:#-]?\s*([A-Za-z0-9/-]+)", remainder, re.IGNORECASE)
        if ref_match:
            reference = ref_match.group(1)

        return GenericTransaction(
            date=date_text,
            description=description,
            debit=float(debit or 0.0),
            credit=float(credit or 0.0),
            balance=balance,
            reference=reference,
            raw_line=text,
        )

    def _normalize_date(self, date_text: str) -> str:
        candidates = [
            "%d-%m-%Y",
            "%d/%m/%Y",
            "%d-%m-%y",
            "%d/%m/%y",
            "%d %b %Y",
            "%d %b %y",
            "%b %d %Y",
            "%b %d, %Y",
        ]
        normalized = date_text.replace("/", "-").replace(",", "")
        for fmt in candidates:
            try:
                dt = datetime.strptime(normalized, fmt)
                return dt.strftime("%Y-%m-%d")
            except Exception:
                continue
        return date_text

    def _extract_amounts(self, text: str) -> List[float]:
        amounts: List[float] = []
        for match in re.finditer(r"\(?-?\d{1,3}(?:,\d{3})*(?:\.\d{1,2})?\)?|\(?-?\d+(?:\.\d{1,2})?\)?", text):
            value = self._clean_amount(match.group(0))
            if value is not None:
                amounts.append(value)
        return amounts

    def _clean_amount(self, value: str) -> Optional[float]:
        if not value:
            return None
        try:
            cleaned = value.replace(",", "").replace("(", "-").replace(")", "")
            return float(cleaned)
        except Exception:
            return None

    def _clean_description(self, text: str) -> str:
        text = re.sub(r"\b(?:ref|utr|rrn|txn|transaction)\s*[:#-]?\s*[A-Za-z0-9/-]+", " ", text, flags=re.IGNORECASE)
        text = re.sub(r"\(?-?\d{1,3}(?:,\d{3})*(?:\.\d{1,2})?\)?", " ", text)
        text = re.sub(r"\s+", " ", text)
        return text.strip(" -|\t")


class GenericTransactionValidator:
    def __init__(self, strict_mode: bool = True):
        self.strict_mode = strict_mode

    def validate(self, transactions: List[Dict[str, Any]]) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
        validated: List[Dict[str, Any]] = []
        errors = 0
        for idx, txn in enumerate(transactions, start=1):
            try:
                validated.append(self._validate_one(txn, idx))
            except GenericValidationError:
                errors += 1
                if self.strict_mode:
                    raise
        return validated, {"validated": len(validated), "errors": errors}

    def _validate_one(self, txn: Dict[str, Any], idx: int) -> Dict[str, Any]:
        if not isinstance(txn, dict):
            raise GenericValidationError("Transaction must be a dict", error_code="INVALID_TRANSACTION")
        date = str(txn.get("date") or "").strip()
        description = str(txn.get("description") or txn.get("narration") or "").strip()
        if not date:
            raise GenericValidationError(f"Transaction {idx} missing date", error_code="MISSING_DATE")
        if not description:
            raise GenericValidationError(f"Transaction {idx} missing description", error_code="MISSING_DESCRIPTION")
        debit = self._to_float(txn.get("debit"))
        credit = self._to_float(txn.get("credit"))
        balance = self._to_float(txn.get("balance"), allow_none=True)
        if debit > 0 and credit > 0:
            if self.strict_mode:
                raise GenericValidationError(f"Transaction {idx} has both debit and credit", error_code="BOTH_DEBIT_CREDIT")
        txn = dict(txn)
        txn["date"] = date
        txn["description"] = description
        txn["debit"] = debit
        txn["credit"] = credit
        txn["balance"] = balance
        return txn

    @staticmethod
    def _to_float(value: Any, allow_none: bool = False) -> float:
        if value is None or value == "":
            return 0.0 if not allow_none else None
        try:
            return float(str(value).replace(",", ""))
        except Exception:
            return 0.0 if not allow_none else None


class GenericReconciliation:
    def __init__(self, strict_mode: bool = True):
        self.strict_mode = strict_mode

    def reconcile(
        self,
        transactions: List[Dict[str, Any]],
        expected_opening: Optional[float] = None,
        expected_closing: Optional[float] = None,
    ) -> Dict[str, Any]:
        if not transactions:
            return {"passed": True, "message": "No transactions to reconcile", "opening_balance": expected_opening, "closing_balance": expected_closing}

        opening = expected_opening
        closing = expected_closing
        if opening is None:
            opening = self._first_balance(transactions)
        if closing is None:
            closing = self._last_balance(transactions)

        if opening is None or closing is None:
            return {"passed": not self.strict_mode, "message": "Insufficient balance data", "opening_balance": opening, "closing_balance": closing}

        computed = opening
        for txn in transactions:
            computed += float(txn.get("credit") or 0) - float(txn.get("debit") or 0)
        passed = abs(computed - closing) < 1.0
        return {"passed": passed, "message": "Reconciliation complete", "opening_balance": opening, "closing_balance": closing, "computed_closing": computed}

    @staticmethod
    def _first_balance(transactions: List[Dict[str, Any]]) -> Optional[float]:
        for txn in transactions:
            bal = txn.get("balance")
            if bal is not None:
                return float(bal)
        return None

    @staticmethod
    def _last_balance(transactions: List[Dict[str, Any]]) -> Optional[float]:
        for txn in reversed(transactions):
            bal = txn.get("balance")
            if bal is not None:
                return float(bal)
        return None


class GenericRecurringEngine:
    def detect(self, transactions: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        counts: Dict[str, int] = {}
        for txn in transactions:
            key = self._normalize_key(str(txn.get("description") or ""))
            counts[key] = counts.get(key, 0) + 1
        for txn in transactions:
            key = self._normalize_key(str(txn.get("description") or ""))
            txn["is_recurring"] = counts.get(key, 0) > 1
        return transactions

    @staticmethod
    def _normalize_key(text: str) -> str:
        text = re.sub(r"\d+", "", text.lower())
        text = re.sub(r"\s+", " ", text)
        return text.strip()


class GenericAggregationEngine:
    def aggregate(
        self,
        transactions: List[Dict[str, Any]],
        opening: Optional[float] = None,
        closing: Optional[float] = None,
    ) -> Dict[str, Any]:
        category_totals: Dict[str, Dict[str, float]] = {}
        monthly_totals: Dict[str, Dict[str, float]] = {}
        recurring_count = 0
        for txn in transactions:
            category = str(txn.get("category") or "Others")
            debit = float(txn.get("debit") or 0)
            credit = float(txn.get("credit") or 0)
            balance = txn.get("balance")
            category_totals.setdefault(category, {"debit": 0.0, "credit": 0.0, "count": 0})
            category_totals[category]["debit"] += debit
            category_totals[category]["credit"] += credit
            category_totals[category]["count"] += 1
            date = str(txn.get("date") or "")
            month = date[:7] if len(date) >= 7 else "unknown"
            monthly_totals.setdefault(month, {"debit": 0.0, "credit": 0.0})
            monthly_totals[month]["debit"] += debit
            monthly_totals[month]["credit"] += credit
            if txn.get("is_recurring"):
                recurring_count += 1
        return {
            "opening_balance": opening,
            "closing_balance": closing,
            "total_transactions": len(transactions),
            "recurring_count": recurring_count,
            "categories": category_totals,
            "monthly": monthly_totals,
        }


def _aggregation_value(aggregation: Any, key: str, default: Any = None) -> Any:
    if aggregation is None:
        return default
    if isinstance(aggregation, dict):
        return aggregation.get(key, default)
    getter = getattr(aggregation, "get", None)
    if callable(getter):
        try:
            return getter(key, default)
        except TypeError:
            pass
    return getattr(aggregation, key, default)


class GenericExcelGenerator:
    def __init__(self, bank_config: GenericBankConfig):
        self.bank_config = bank_config

    def generate(self, transactions: List[Dict[str, Any]], aggregation: Dict[str, Any], user_info: Dict[str, Any], output_path: str) -> str:
        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"
        self._build_summary_sheet(ws, aggregation, user_info)
        self._build_transactions_sheet(wb.create_sheet("Transactions"), transactions)
        self._build_categories_sheet(wb.create_sheet("Categories"), aggregation)
        self._build_monthly_sheet(wb.create_sheet("Monthly"), aggregation)
        self._build_recurring_sheet(wb.create_sheet("Recurring"), transactions)
        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
        wb.save(output_path)
        return output_path

    def _build_summary_sheet(self, ws, aggregation: Dict[str, Any], user_info: Dict[str, Any]) -> None:
        ws["A1"] = f"{self.bank_config.bank_name} Statement Summary"
        ws["A1"].font = Font(bold=True, size=14)
        rows = [
            ("Account Holder", user_info.get("full_name", "")),
            ("Bank", self.bank_config.bank_name),
            ("Opening Balance", _aggregation_value(aggregation, "opening_balance")),
            ("Closing Balance", _aggregation_value(aggregation, "closing_balance")),
            ("Total Transactions", _aggregation_value(aggregation, "total_transactions", 0)),
            ("Recurring Count", _aggregation_value(aggregation, "recurring_count", 0)),
        ]
        for idx, (label, value) in enumerate(rows, start=3):
            ws[f"A{idx}"] = label
            ws[f"B{idx}"] = value
            ws[f"A{idx}"].font = Font(bold=True)

    def _build_transactions_sheet(self, ws, transactions: List[Dict[str, Any]]) -> None:
        headers = ["Date", "Description", "Debit", "Credit", "Balance", "Category", "Confidence", "Recurring"]
        ws.append(headers)
        for row in transactions:
            ws.append([
                row.get("date"), row.get("description"), row.get("debit"), row.get("credit"),
                row.get("balance"), row.get("category"), row.get("confidence"), row.get("is_recurring"),
            ])
        self._style_header(ws)

    def _build_categories_sheet(self, ws, aggregation: Dict[str, Any]) -> None:
        ws.append(["Category", "Debit", "Credit", "Count"])
        for category, values in _aggregation_value(aggregation, "categories", {}).items():
            ws.append([category, values.get("debit", 0), values.get("credit", 0), values.get("count", 0)])
        self._style_header(ws)

    def _build_monthly_sheet(self, ws, aggregation: Dict[str, Any]) -> None:
        ws.append(["Month", "Debit", "Credit"])
        for month, values in _aggregation_value(aggregation, "monthly", {}).items():
            ws.append([month, values.get("debit", 0), values.get("credit", 0)])
        self._style_header(ws)

    def _build_recurring_sheet(self, ws, transactions: List[Dict[str, Any]]) -> None:
        ws.append(["Date", "Description", "Recurring"])
        for row in transactions:
            if row.get("is_recurring"):
                ws.append([row.get("date"), row.get("description"), True])
        self._style_header(ws)

    @staticmethod
    def _style_header(ws) -> None:
        fill = PatternFill("solid", fgColor="1F2937")
        font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = min(max(length + 2, 12), 40)


class GenericAIFallback:
    def __init__(self, bank_config: GenericBankConfig, api_key: Optional[str] = None):
        self.bank_config = bank_config
        self.api_key = api_key

    def classify(self, transactions: List[Dict[str, Any]], bank_name: str, account_type: str) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
        for txn in transactions:
            if not txn.get("category"):
                txn["category"] = "Others"
            txn["source"] = txn.get("source") or "ai_fallback"
        return transactions, {"provider": "generic", "count": len(transactions)}


class GenericRuleEngine:
    def __init__(self, bank_config: GenericBankConfig, keywords_file: Optional[str] = None):
        self.classifier = GenericClassifier(bank_config, keywords_file=keywords_file)

    def classify(self, transactions: List[Dict[str, Any]]) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
        return self.classifier.classify(transactions)


class GenericProcessor:
    def __init__(
        self,
        bank_config: GenericBankConfig,
        strict_mode: bool = True,
        enable_ai: bool = False,
        api_key: Optional[str] = None,
    ):
        self.bank_config = bank_config
        self.strict_mode = strict_mode
        self.enable_ai = enable_ai
        self.api_key = api_key
        self.structure_validator = GenericStructureValidator(bank_config)
        self.parser = GenericParser(bank_config)
        self.transaction_validator = GenericTransactionValidator(strict_mode=False)
        self.reconciliation = GenericReconciliation(strict_mode=False)
        self.rule_engine = GenericRuleEngine(bank_config)
        self.ai_fallback = GenericAIFallback(bank_config, api_key=api_key)
        self.recurring_engine = GenericRecurringEngine()
        self.aggregation_engine = GenericAggregationEngine()
        self.excel_generator = GenericExcelGenerator(bank_config)
        self.logger = logging.getLogger(f"{__name__}.{self.__class__.__name__}")

    def process(self, file_path: str, user_info: Dict[str, Any], output_dir: Optional[str] = None) -> GenericProcessingResult:
        pipeline_start = time.monotonic()
        metrics = GenericProcessingMetrics()
        try:
            step = time.monotonic()
            parse_result = self.parser.parse(file_path)
            metrics.step_timings["parsing"] = round((time.monotonic() - step) * 1000, 1)

            transactions = [txn.to_dict() for txn in parse_result.transactions]
            validated, validation_stats = self.transaction_validator.validate(transactions)
            metrics.step_timings["validation"] = validation_stats.get("validated", 0)
            metrics.transaction_count = len(validated)

            opening = parse_result.metadata.get("opening_balance") if parse_result.metadata else None
            closing = parse_result.metadata.get("closing_balance") if parse_result.metadata else None
            reconciliation = self.reconciliation.reconcile(validated, opening, closing)
            metrics.reconciliation_passed = bool(reconciliation.get("passed"))
            metrics.step_timings["reconciliation"] = round(0.0, 1)

            classified, unclassified = self.rule_engine.classify(validated)
            combined = classified + unclassified
            metrics.classified_count = len(classified)
            metrics.unclassified_count = len(unclassified)

            combined = self.recurring_engine.detect(combined)
            metrics.recurring_count = sum(1 for txn in combined if txn.get("is_recurring"))

            aggregation = self.aggregation_engine.aggregate(combined, opening, closing)

            if output_dir is None:
                output_dir = os.path.dirname(file_path) or "."
            excel_filename = f"{self.bank_config.file_prefix}_{uuid.uuid4().hex[:12]}.xlsx"
            excel_path = os.path.join(output_dir, excel_filename)
            step = time.monotonic()
            self.excel_generator.generate(combined, aggregation, user_info, excel_path)
            metrics.step_timings["report_generation"] = round((time.monotonic() - step) * 1000, 1)
            metrics.total_time_ms = round((time.monotonic() - pipeline_start) * 1000, 1)

            return GenericProcessingResult(
                status="success",
                excel_path=excel_path,
                transactions=combined,
                aggregation=aggregation,
                metrics=metrics,
                bank_key=self.bank_config.bank_key,
            )
        except Exception as exc:
            self.logger.error("%s processing failed: %s", self.bank_config.bank_name, str(exc), exc_info=True)
            raise GenericProcessorError(
                f"{self.bank_config.bank_name} processing failed: {str(exc)}",
                stage="processing",
                error_code=f"{self.bank_config.bank_key.upper()}_PROCESSING_ERROR",
            )


def generate_report(transactions: List[Dict[str, Any]], output_path: str, user_info: Dict[str, Any], bank_config: GenericBankConfig) -> Dict[str, Any]:
    aggregator = GenericAggregationEngine()
    excel = GenericExcelGenerator(bank_config)
    aggregation = aggregator.aggregate(transactions, user_info.get("opening_balance"), user_info.get("closing_balance"))
    excel.generate(transactions, aggregation, user_info, output_path)
    return {
        "total_transactions": len(transactions),
        "recurring_count": _aggregation_value(aggregation, "recurring_count", 0),
        "categories_used": len(_aggregation_value(aggregation, "categories", {})),
        "output_path": output_path,
    }
