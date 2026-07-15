import logging
import os
from typing import Any, Dict, Optional

logger = logging.getLogger(__name__)

# Sheet keys aligned to Lite Excel sheet order (9 sheets).
# Credit Card Payments comes before Transactions (last).
SHEET_KEYS = [
    "account_summary",           # 1 Summary
    "monthly_analysis",          # 2 Monthly Analysis
    "top5_credits",              # 3 Top 5 Credits
    "top5_debits",               # 4 Top 5 Debits
    "bounces_penal",             # 5 Bounce And Penal Transactions
    "salary_transactions",       # 6 Salary Transactions
    "loan_repayment",            # 7 Loan Repayment Transactions
    "credit_card_payments",      # 8 Credit Card Payments
    "raw_transactions",          # 9 Transactions (always last)
]

MAX_PREVIEW_ROWS = 50


def extract_sheet_previews(excel_path: str) -> Dict[str, Dict[str, Any]]:
    """Extract sheet preview data for frontend consumers."""
    previews: Dict[str, Dict[str, Any]] = {}
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(excel_path, read_only=True, data_only=True)

        for idx, sheet_name in enumerate(workbook.sheetnames):
            if idx >= len(SHEET_KEYS):
                break

            worksheet = workbook[sheet_name]
            rows_data = []
            headers = []
            pending_title = None

            for row_idx, row in enumerate(worksheet.iter_rows(values_only=True)):
                string_row = [str(cell) if cell is not None else "" for cell in row]

                is_single_title_row = sum(1 for cell in string_row if cell.strip()) <= 1
                if row_idx == 0 and is_single_title_row:
                    pending_title = string_row[0] if string_row and string_row[0].strip() else sheet_name
                    continue

                looks_like_header_row = (
                    len(string_row) >= 2
                    and string_row[0].strip().lower() == "category"
                    and string_row[1].strip().lower() == "source"
                )
                if headers and looks_like_header_row:
                    continue

                if is_single_title_row:
                    continue

                if not headers:
                    headers = string_row
                    if pending_title and headers and headers[0] == pending_title:
                        continue
                    continue

                if row_idx <= MAX_PREVIEW_ROWS + 1:
                    rows_data.append(string_row)

            previews[SHEET_KEYS[idx]] = {
                "title": f"Sheet {idx + 1} - {sheet_name}",
                "headers": headers,
                "rows": rows_data,
            }

        workbook.close()
    except Exception as exc:
        logger.warning("Could not extract sheet previews: %s", str(exc))

    return previews


def build_frontend_processing_result(
    result: Dict[str, Any],
    mode: str,
    excel_url: Optional[str] = None,
) -> Dict[str, Any]:
    """Transform processing output into the shape expected by the frontend result view."""
    excel_path = result.get("excel_path", "")

    resolved_excel_url = excel_url or ""
    if not resolved_excel_url and excel_path:
        excel_filename = os.path.basename(excel_path)
        resolved_excel_url = f"/download/{excel_filename}" if excel_filename else ""

    previews: Dict[str, Dict[str, Any]] = {}
    if excel_path and os.path.isfile(excel_path):
        previews = extract_sheet_previews(excel_path)

    response: Dict[str, Any] = {
        "status": result.get("status", "success"),
        "mode": mode or result.get("mode", "free"),
        "bank": result.get("bank_key", ""),
        "excel_path": excel_path,
        "excel_url": resolved_excel_url,
        "pdf_url": result.get("pdf_url", ""),
        "statement_profile": result.get("statement_profile", {}),
        "financial_profile": result.get("financial_profile", {}),
        "stats": result.get("stats", {}),
        "validation": result.get("validation", {}),
        "performance": result.get("performance", {}),
        "bank_key": result.get("bank_key"),
        "excel_object_key": result.get("excel_object_key"),
        "source_pdf_object_key": result.get("source_pdf_object_key"),
        **previews,
    }

    if result.get("ai_usage") is not None:
        response["ai_usage"] = result.get("ai_usage")

    if result.get("error"):
        response["error"] = result["error"]

    return response
