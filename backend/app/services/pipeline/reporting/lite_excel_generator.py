"""
Airco Insights Lite — shared 9-sheet Excel generator.

Sheet order (exact):
1. Summary
2. Monthly Analysis
3. Top 5 Credits
4. Top 5 Debits
5. Bounce And Penal Transactions
6. Salary Transactions
7. Loan Repayment Transactions
8. Transactions
9. Credit Card Payments
"""

from __future__ import annotations

import logging
import os
from typing import Any, Dict, List, Optional, Sequence, Union

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.services.pipeline.reporting.lite_metrics import (
    ACCOUNT_INFO_KEYS,
    MONTHLY_METRIC_KEYS,
    SUMMARY_STAT_KEYS,
    build_lite_report_model,
)

logger = logging.getLogger(__name__)

SHEET_ORDER = [
    "Summary",
    "Monthly Analysis",
    "Top 5 Credits",
    "Top 5 Debits",
    "Bounce And Penal Transactions",
    "Salary Transactions",
    "Loan Repayment Transactions",
    "Credit Card Payments",
    "Transactions",
]

# All Account Statistics rows are month-wise (one column per month).
PER_MONTH_SUMMARY_KEYS = set(SUMMARY_STAT_KEYS)

# Integer-style metrics (counts) — no decimal places in Summary.
COUNT_SUMMARY_KEYS = {
    "Total Credit Count",
    "Total Debit Count",
    "Total Cheque",
    "Cnt of Cheque Bounces",
}

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
SECTION_FILL = PatternFill("solid", fgColor="D6DCE4")
SECTION_FONT = Font(name="Calibri", bold=True, size=12, color="1F4E79")
LABEL_FONT = Font(name="Calibri", bold=True, size=10)
VALUE_FONT = Font(name="Calibri", size=10)
MONTH_HEADER_FILL = PatternFill("solid", fgColor="2E75B6")
ALT_ROW_FILL = PatternFill("solid", fgColor="F2F2F2")
THIN = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)


def _style_header_row(ws, row: int, col_count: int) -> None:
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN
    ws.freeze_panes = ws.cell(row=row + 1, column=1)
    ws.auto_filter.ref = f"A{row}:{get_column_letter(col_count)}{row}"


def _autosize(ws, min_width: int = 12, max_width: int = 42) -> None:
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        length = 0
        for cell in col_cells:
            if cell.value is None:
                continue
            length = max(length, len(str(cell.value)))
        ws.column_dimensions[letter].width = max(min_width, min(max_width, length + 2))


def _write_kv(ws, row: int, label: str, value: Any) -> None:
    label_cell = ws.cell(row=row, column=1, value=label)
    label_cell.font = LABEL_FONT
    label_cell.border = THIN
    # Always keep the row; missing values render as blank (not None / not fabricated).
    if value is None or value == "":
        display: Any = ""
    else:
        display = value
    value_cell = ws.cell(row=row, column=2, value=display if display != "" else None)
    # Force blank string into cell.data_type for string fields so Excel shows empty not 0
    if display == "":
        value_cell.value = ""
    value_cell.font = VALUE_FONT
    value_cell.border = THIN
    if isinstance(display, (int, float)) and not isinstance(display, bool):
        value_cell.number_format = "#,##0.00" if isinstance(display, float) else "0"
        value_cell.alignment = Alignment(horizontal="right")


def _write_summary_sheet(ws, model: Dict[str, Any]) -> None:
    account_info = model["account_info"]
    stats = model["summary_stats"]
    months: List[str] = list(model["month_labels"] or [])

    ws.cell(row=1, column=1, value="Account Information").font = SECTION_FONT
    ws.cell(row=1, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)

    row = 2
    for key in ACCOUNT_INFO_KEYS:
        _write_kv(ws, row, key, account_info.get(key, ""))
        row += 1

    row += 1
    end_col = max(2, 1 + len(months))
    section_cell = ws.cell(row=row, column=1, value="Account Statistics")
    section_cell.font = SECTION_FONT
    section_cell.fill = SECTION_FILL
    for col in range(1, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = SECTION_FILL
        cell.border = THIN
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
    row += 1

    # Header row: Metric | Month-1 | Month-2 | ...
    if months:
        metric_hdr = ws.cell(row=row, column=1, value="Metric")
        metric_hdr.font = HEADER_FONT
        metric_hdr.fill = HEADER_FILL
        metric_hdr.border = THIN
        metric_hdr.alignment = Alignment(horizontal="center", vertical="center")
        for idx, month in enumerate(months, start=2):
            cell = ws.cell(row=row, column=idx, value=month)
            cell.font = HEADER_FONT
            cell.fill = MONTH_HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN
        header_row = row
        row += 1
    else:
        header_row = None

    for r_offset, key in enumerate(SUMMARY_STAT_KEYS):
        value = stats.get(key, "")
        label_cell = ws.cell(row=row, column=1, value=key)
        label_cell.font = LABEL_FONT
        label_cell.border = THIN
        if r_offset % 2 == 1:
            label_cell.fill = ALT_ROW_FILL

        month_map = value if isinstance(value, dict) else {}
        is_count = key in COUNT_SUMMARY_KEYS
        is_pct = "%" in key

        if months:
            for idx, month in enumerate(months, start=2):
                raw = month_map.get(month, 0.0)
                if raw is None or raw == "":
                    raw = 0.0
                # Counts stay whole numbers; amounts/pct keep decimals.
                if is_count:
                    display: Any = int(round(float(raw)))
                    num_fmt = "0"
                else:
                    display = float(raw)
                    num_fmt = '0.00"%"' if is_pct else "#,##0.00"
                cell = ws.cell(row=row, column=idx, value=display)
                cell.font = VALUE_FONT
                cell.border = THIN
                cell.number_format = num_fmt
                cell.alignment = Alignment(horizontal="right")
                if r_offset % 2 == 1:
                    cell.fill = ALT_ROW_FILL
        else:
            cell = ws.cell(row=row, column=2, value=0)
            cell.font = VALUE_FONT
            cell.border = THIN
            cell.number_format = "0" if is_count else "#,##0.00"
        row += 1

    if header_row:
        ws.freeze_panes = ws.cell(row=header_row + 1, column=2)
    _autosize(ws)
    # Wider first column for metric labels
    ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width or 12, 26)


def _write_monthly_sheet(ws, model: Dict[str, Any]) -> None:
    months: List[str] = list(model["month_labels"] or [])
    monthly = model["monthly_analysis"]

    headers = ["Metric", *months] if months else ["Metric", "Value"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        if col == 1:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        else:
            cell.fill = MONTH_HEADER_FILL
            cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN
    ws.freeze_panes = "B2"
    if headers:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    for r_offset, metric in enumerate(MONTHLY_METRIC_KEYS):
        r_idx = r_offset + 2
        label_cell = ws.cell(row=r_idx, column=1, value=metric)
        label_cell.font = LABEL_FONT
        label_cell.border = THIN
        if r_offset % 2 == 1:
            label_cell.fill = ALT_ROW_FILL
        values = monthly.get(metric, {})
        is_count = metric.endswith("Count") or metric in {
            "inwBounce",
            "owtBounce",
            "creditCount",
            "debitCount",
        }
        if months:
            for c_idx, month in enumerate(months, start=2):
                val = values.get(month, 0.0)
                if is_count:
                    display: Any = int(round(float(val or 0)))
                    num_fmt = "0"
                else:
                    display = float(val or 0)
                    num_fmt = "#,##0.00"
                cell = ws.cell(row=r_idx, column=c_idx, value=display)
                cell.border = THIN
                cell.font = VALUE_FONT
                cell.number_format = num_fmt
                cell.alignment = Alignment(horizontal="right")
                if r_offset % 2 == 1:
                    cell.fill = ALT_ROW_FILL
        else:
            cell = ws.cell(row=r_idx, column=2, value=0)
            cell.border = THIN
    _autosize(ws)
    ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width or 12, 22)


def _write_top_sheet(ws, rows: Sequence[Dict[str, Any]], amount_key: str) -> None:
    headers = ["Date", "Description", "Category", "Amount", "Balance", "Month"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
    _style_header_row(ws, 1, len(headers))

    if not rows:
        ws.cell(row=2, column=1, value="No transactions found")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
        _autosize(ws)
        return

    for r_idx, row in enumerate(rows, start=2):
        values = [
            row.get("date_display", ""),
            row.get("description", ""),
            row.get("category", ""),
            row.get(amount_key, 0.0),
            row.get("balance", 0.0),
            row.get("month", ""),
        ]
        for c_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = THIN
            cell.font = VALUE_FONT
            if c_idx in {4, 5} and isinstance(value, (int, float)):
                cell.number_format = "#,##0.00"
    _autosize(ws)


def _write_txn_sheet(
    ws,
    rows: Sequence[Dict[str, Any]],
    *,
    include_ref: bool = False,
) -> None:
    if include_ref:
        headers = ["Date", "Description", "Ref/Cheque No", "Debit", "Credit", "Balance", "Category", "Month"]
    else:
        headers = ["Date", "Description", "Category", "Debit", "Credit", "Balance", "Month"]

    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
    _style_header_row(ws, 1, len(headers))

    if not rows:
        ws.cell(row=2, column=1, value="No transactions found")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
        _autosize(ws)
        return

    for r_idx, row in enumerate(rows, start=2):
        if include_ref:
            values = [
                row.get("date_display", ""),
                row.get("description", ""),
                row.get("ref_no", ""),
                row.get("debit", 0.0),
                row.get("credit", 0.0),
                row.get("balance", 0.0),
                row.get("category", ""),
                row.get("month", ""),
            ]
            money_cols = {4, 5, 6}
        else:
            values = [
                row.get("date_display", ""),
                row.get("description", ""),
                row.get("category", ""),
                row.get("debit", 0.0),
                row.get("credit", 0.0),
                row.get("balance", 0.0),
                row.get("month", ""),
            ]
            money_cols = {4, 5, 6}

        for c_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = THIN
            cell.font = VALUE_FONT
            if c_idx in money_cols and isinstance(value, (int, float)):
                cell.number_format = "#,##0.00"
    _autosize(ws)


class LiteExcelGenerator:
    """Shared Lite Excel generator used by all banks."""

    def generate(
        self,
        transactions: Sequence[Any],
        metadata: Optional[Dict[str, Any]] = None,
        output_path: str = "",
    ) -> str:
        if not output_path:
            raise ValueError("output_path is required")

        parent = os.path.dirname(output_path)
        if parent:
            os.makedirs(parent, exist_ok=True)

        model = build_lite_report_model(transactions, metadata or {})
        wb = Workbook()

        # openpyxl creates one default sheet; rename/reuse it for Summary
        default = wb.active
        default.title = "Summary"
        _write_summary_sheet(default, model)

        monthly_ws = wb.create_sheet("Monthly Analysis")
        _write_monthly_sheet(monthly_ws, model)

        top_credit_ws = wb.create_sheet("Top 5 Credits")
        _write_top_sheet(top_credit_ws, model["top5_credits"], "credit")

        top_debit_ws = wb.create_sheet("Top 5 Debits")
        _write_top_sheet(top_debit_ws, model["top5_debits"], "debit")

        bounce_ws = wb.create_sheet("Bounce And Penal Transactions")
        _write_txn_sheet(bounce_ws, model["bounce_penal"], include_ref=False)

        salary_ws = wb.create_sheet("Salary Transactions")
        _write_txn_sheet(salary_ws, model["salary"], include_ref=False)

        loan_ws = wb.create_sheet("Loan Repayment Transactions")
        _write_txn_sheet(loan_ws, model["loan"], include_ref=False)

        # Credit Card Payments before Transactions (Transactions is always last)
        cc_ws = wb.create_sheet("Credit Card Payments")
        _write_txn_sheet(cc_ws, model["credit_card"], include_ref=False)

        all_ws = wb.create_sheet("Transactions")
        _write_txn_sheet(all_ws, model["transactions"], include_ref=True)

        # Enforce exact order
        for idx, name in enumerate(SHEET_ORDER):
            wb.move_sheet(name, offset=idx - wb.sheetnames.index(name))

        wb.save(output_path)
        logger.info(
            "LiteExcelGenerator: wrote %d sheets / %d transactions -> %s",
            len(wb.sheetnames),
            len(model["transactions"]),
            output_path,
        )
        return output_path


def generate_lite_excel(
    transactions: Sequence[Any],
    metadata: Optional[Dict[str, Any]] = None,
    output_path: str = "",
) -> str:
    return LiteExcelGenerator().generate(transactions, metadata, output_path)
