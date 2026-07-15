import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from app.services.banks.canara.processor import CanaraProcessor


SAMPLES_DIR = Path(__file__).resolve().parents[3] / "samples" / "New Bank Samples" / "Canara"


def _iso_date(value):
    if value is None:
        return None
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d")
    return str(value)[:10]


@pytest.mark.parametrize(
    "pdf_name, expected_count, expected_first_credit, expected_from, expected_to",
    [
        (
            "68d4de06cbe50_1758783422156.pdf",
            419,
            105284.0,
            "2025-03-01",
            "2025-09-25",
        ),
    ],
)
def test_canara_first_row_credit_and_summary_dates(pdf_name, expected_count, expected_first_credit, expected_from, expected_to, tmp_path):
    pdf_path = SAMPLES_DIR / pdf_name
    processor = CanaraProcessor(strict_mode=False, enable_ai=False)

    result = processor.process(
        str(pdf_path),
        {
            "bank_name": "Canara Bank",
            "full_name": "",
            "account_type": "salaried",
            "account_no": "",
        },
        str(tmp_path),
    )

    assert result.status == "success"
    assert len(result.transactions) == expected_count
    assert result.transactions[0]["debit"] is None
    assert result.transactions[0]["credit"] == pytest.approx(expected_first_credit)

    wb = load_workbook(result.excel_path, data_only=True, read_only=True)
    try:
        raw_ws = wb["Raw Transaction"]
        raw_rows = [row for row in raw_ws.iter_rows(min_row=2, values_only=True) if row and row[0]]
        assert len(raw_rows) == expected_count
        assert raw_rows[0][2] in (None, "")
        assert raw_rows[0][3] == pytest.approx(expected_first_credit)

        summary_ws = wb["Summary"]
        assert _iso_date(summary_ws["B3"].value) == expected_from
        assert _iso_date(summary_ws["B4"].value) == expected_to
    finally:
        wb.close()
