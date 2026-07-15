import sys
from collections import Counter
from pathlib import Path

import pytest
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from app.services.banks.icici.processor import ICICIProcessor


SAMPLES_DIR = Path(__file__).resolve().parents[3] / "samples" / "New Bank Samples" / "ICIC"


@pytest.mark.parametrize(
    "pdf_name, expected_count, expected_01_10_count, expected_eenadutele_amount",
    [
        (
            "Statement__1760079678114.pdf",
            196,
            8,
            4501.0,
        ),
    ],
)
def test_icici_parser_keeps_eenadutele_mandate_row(pdf_name, expected_count, expected_01_10_count, expected_eenadutele_amount, tmp_path):
    pdf_path = SAMPLES_DIR / pdf_name
    processor = ICICIProcessor(strict_mode=False, enable_ai=False)

    result = processor.process(
        str(pdf_path),
        {
            "bank_name": "ICICI Bank",
            "full_name": "",
            "account_type": "salaried",
            "account_no": "",
        },
        str(tmp_path),
    )

    assert result.status == "success"
    assert len(result.transactions) == expected_count

    date_counts = Counter(txn["date"] for txn in result.transactions)
    assert date_counts["2025-10-01"] == expected_01_10_count

    eenadutele = [
        txn for txn in result.transactions
        if txn["date"] == "2025-10-01" and "EENADUTELE" in (txn.get("description") or "").upper()
    ]
    assert len(eenadutele) == 1
    assert eenadutele[0]["debit"] == pytest.approx(expected_eenadutele_amount)
    assert eenadutele[0]["credit"] is None

    wb = load_workbook(result.excel_path, data_only=True, read_only=True)
    try:
        raw_ws = wb["Raw Transaction"]
        raw_rows = [row for row in raw_ws.iter_rows(min_row=2, values_only=True) if row and row[0]]
        assert len(raw_rows) == expected_count
        assert sum(1 for row in raw_rows if getattr(row[0], "date", None) and row[0].date().isoformat() == "2025-10-01") == expected_01_10_count
    finally:
        wb.close()
