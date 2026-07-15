"""Unit tests for Airco Insights Lite 9-sheet Excel export (pipeline.reporting)."""

from __future__ import annotations

import os
import tempfile
from decimal import Decimal

from openpyxl import load_workbook

from app.services.pipeline.models import NormalizedTransaction
from app.services.pipeline.reporting import (
    LiteExcelGenerator,
    MONTHLY_METRIC_KEYS,
    SHEET_ORDER,
    build_lite_report_model,
)

SAMPLE_TXNS = [
    {
        "date": "01-07-2023",
        "description": "SALARY CREDIT ACME",
        "debit": 0,
        "credit": 50000,
        "balance": 50000,
        "category": "Salary",
    },
    {
        "date": "05-07-2023",
        "description": "ATM WDL",
        "debit": 2000,
        "credit": 0,
        "balance": 48000,
        "category": "ATM Withdrawal",
    },
    {
        "date": "10-07-2023",
        "description": "CHQ BOUNCE RETURN",
        "debit": 500,
        "credit": 0,
        "balance": 47500,
        "category": "Bank Charges",
    },
    {
        "date": "15-07-2023",
        "description": "EMI LOAN HDFC",
        "debit": 10000,
        "credit": 0,
        "balance": 37500,
        "category": "Loan Payment",
    },
    {
        "date": "20-07-2023",
        "description": "CC PAYMENT VISA",
        "debit": 3000,
        "credit": 0,
        "balance": 34500,
        "category": "Credit Card Payment",
    },
    {
        "date": "01-08-2023",
        "description": "NEFT CR FRIEND",
        "debit": 0,
        "credit": 15000,
        "balance": 49500,
        "category": "Transfer",
    },
    {
        "date": "12-08-2023",
        "description": "CASH DEPOSIT",
        "debit": 0,
        "credit": 2000,
        "balance": 51500,
        "category": "Business Income",
    },
    {
        "date": "25-08-2023",
        "description": "PENALTY CHARGE",
        "debit": 100,
        "credit": 0,
        "balance": 51400,
        "category": "Bank Charges",
    },
]

SAMPLE_META = {
    "accountName": "MR. VINAY S",
    "account_number": "50100433372120",
    "bank_name": "HDFC",
    "account_type": "SAVING",
    "IFSC": "HDFC0000261",
    "statement_from": "2023-07-01",
    "statement_to": "2023-08-31",
    "email": "vinay.s008@gmail.com",
    "address": "Bangalore",
    "opening_balance": 0,
    "closing_balance": 51400,
    "currentBalance": 51400,
    "account_open_date": "2021-02-01",
}


def test_lite_report_model_cross_checks():
    model = build_lite_report_model(SAMPLE_TXNS, SAMPLE_META)
    assert model["account_info"]["accountName"] == "MR. VINAY S"
    assert model["account_info"]["relationshipWithBank"]
    assert "ECS/NACH" in MONTHLY_METRIC_KEYS
    months = model["month_labels"]
    assert months == ["Jul-2023", "Aug-2023"]

    credit_count = model["summary_stats"]["Total Credit Count"]
    credit_amt = model["summary_stats"]["Total Credit Amount"]
    debit_amt = model["summary_stats"]["Total Debit Amount"]
    assert isinstance(credit_count, dict)
    assert sum(credit_count[m] for m in months) == 3
    assert abs(sum(credit_amt[m] for m in months) - 67000) < 0.01
    assert abs(sum(debit_amt[m] for m in months) - 15600) < 0.01
    assert len(MONTHLY_METRIC_KEYS) == 21

    credit_sum = sum(model["monthly_analysis"]["creditValue"][m] for m in months)
    debit_sum = sum(model["monthly_analysis"]["debitValue"][m] for m in months)
    assert abs(credit_sum - sum(credit_amt[m] for m in months)) < 0.01
    assert abs(debit_sum - sum(debit_amt[m] for m in months)) < 0.01

    for month in months:
        salary = model["monthly_analysis"]["salaryCredits"][month]
        non_salary = model["monthly_analysis"]["nonSalaryCredits"][month]
        credit_value = model["monthly_analysis"]["creditValue"][month]
        assert abs(salary + non_salary - credit_value) < 0.01
        assert month in model["summary_stats"]["Total Credit Count"]
        assert month in model["summary_stats"]["Start of Month Balance"]
        assert month in model["summary_stats"]["Balance on Last Day"]

    assert len(model["salary"]) == 1
    assert len(model["loan"]) == 1
    assert len(model["credit_card"]) == 1
    assert len(model["bounce_penal"]) >= 1


def test_summary_sheet_fills_all_month_columns():
    with tempfile.TemporaryDirectory() as tmp:
        path = os.path.join(tmp, "lite_months.xlsx")
        LiteExcelGenerator().generate(SAMPLE_TXNS, SAMPLE_META, path)
        wb = load_workbook(path)
        ws = wb["Summary"]
        header_row = None
        for r in range(1, 40):
            if ws.cell(r, 1).value == "Metric" and ws.cell(r, 2).value == "Jul-2023":
                header_row = r
                break
        assert header_row is not None
        assert ws.cell(header_row, 3).value == "Aug-2023"

        credit_row = None
        for r in range(header_row + 1, header_row + 30):
            if ws.cell(r, 1).value == "Total Credit Count":
                credit_row = r
                break
        assert credit_row is not None
        assert ws.cell(credit_row, 2).value is not None
        assert ws.cell(credit_row, 3).value is not None
        assert int(ws.cell(credit_row, 2).value) + int(ws.cell(credit_row, 3).value) == 3

        ma = wb["Monthly Analysis"]
        assert ma.cell(1, 2).value == "Jul-2023"
        assert ma.cell(1, 3).value == "Aug-2023"
        assert ma.cell(2, 2).value is not None
        assert ma.cell(2, 3).value is not None


def test_hdfc_header_metadata_extraction():
    from app.services.banks.hdfc.structure_validator import HDFCStructureValidator

    sample = """
    AccountBranch : RAJAM
    Address : GROUNDFLOOR,4-323
    BESIDELICOFINDIA
    SRIKAKULAMROAD
    City : RAJAM
    MS MANTAPUDIPADMAVATHI
    State : ANDHRAPRADESH
    Email : padmahrudayvenkat@gmail.com
    CustID : 50529569
    RAJAM532127 AccountNo : 50100004697340 PRIME
    ANDHRAPRADESH A/COpenDate : 04/01/2014
    JOINTHOLDERS: RTGS/NEFTIFSC: HDFC0002284 MICR:532240152
    StatementFrom : 02/12/2022 To : 25/06/2023
    """
    meta = HDFCStructureValidator()._extract_metadata(sample, sample)
    assert meta.account_number == "50100004697340"
    assert meta.account_holder == "MS MANTAPUDIPADMAVATHI"
    assert meta.email == "padmahrudayvenkat@gmail.com"
    assert meta.ifsc == "HDFC0002284"
    assert meta.account_open_date == "04/01/2014"
    assert meta.account_type == "PRIME"
    assert meta.statement_from == "02/12/2022"
    assert meta.address and "RAJAM" in meta.address


def test_lite_report_model_accepts_normalized_transactions():
    rows = [
        NormalizedTransaction(
            date="2023-07-01",
            description="SALARY CREDIT ACME",
            debit=None,
            credit=Decimal("50000"),
            balance=Decimal("50000"),
            category="Salary",
        ),
        NormalizedTransaction(
            date="2023-07-05",
            description="ATM WDL",
            debit=Decimal("2000"),
            credit=None,
            balance=Decimal("48000"),
            category="ATM Withdrawal",
        ),
    ]
    model = build_lite_report_model(rows, SAMPLE_META)
    months = model["month_labels"]
    assert sum(model["summary_stats"]["Total Credit Count"][m] for m in months) == 1
    assert abs(sum(model["summary_stats"]["Total Credit Amount"][m] for m in months) - 50000) < 0.01
    assert abs(sum(model["summary_stats"]["Total Debit Amount"][m] for m in months) - 2000) < 0.01
    assert "Jul-2023" in model["summary_stats"]["Total Credit Count"]
    assert model["summary_stats"]["Total Credit Count"]["Jul-2023"] == 1.0


def test_lite_workbook_has_exactly_nine_sheets():
    with tempfile.TemporaryDirectory() as tmp:
        path = os.path.join(tmp, "lite.xlsx")
        LiteExcelGenerator().generate(SAMPLE_TXNS, SAMPLE_META, path)
        wb = load_workbook(path)
        assert wb.sheetnames == SHEET_ORDER


def test_empty_filtered_sheets_still_created():
    with tempfile.TemporaryDirectory() as tmp:
        path = os.path.join(tmp, "empty.xlsx")
        LiteExcelGenerator().generate([], SAMPLE_META, path)
        wb = load_workbook(path)
        assert wb.sheetnames == SHEET_ORDER
        assert wb["Salary Transactions"]["A2"].value == "No transactions found"
        assert wb["Top 5 Credits"]["A2"].value == "No transactions found"


def test_shim_imports_still_work():
    from app.services.pipeline.reporting.lite_excel_generator import (
        LiteExcelGenerator as ShimGen,
        SHEET_ORDER as ShimOrder,
    )
    from app.services.banks._shared.lite_metrics import build_lite_report_model as shim_model

    assert ShimOrder == SHEET_ORDER
    model = shim_model(SAMPLE_TXNS, SAMPLE_META)
    months = model["month_labels"]
    assert sum(model["summary_stats"]["Total Credit Count"][m] for m in months) == 3
    assert ShimGen is not None
