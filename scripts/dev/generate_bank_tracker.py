from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

wb = Workbook()
wb.remove(wb.active)

# Sheet 1: Executive Summary
sheet1 = wb.create_sheet("Executive Summary")

# Headers
sheet1['A1'] = "AIRCO INSIGHTS - BANK COVERAGE TRACKER"
sheet1['A1'].font = Font(bold=True, size=18, color="FFFFFF")
sheet1['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
sheet1['A1'].alignment = Alignment(horizontal="center")
sheet1.merge_cells('A1:F1')

sheet1['A2'] = "Last Updated:"
sheet1['B2'] = datetime.now().strftime("%Y-%m-%d")
sheet1['B2'].font = Font(bold=True, color="0000FF")
sheet1['A3'] = "Sample PDFs Available:"
sheet1['B3'] = 29
sheet1['B3'].font = Font(bold=True, color="0000FF")

# Key Metrics Section
sheet1['A5'] = "KEY METRICS"
sheet1['A5'].font = Font(bold=True, size=14)
sheet1['A5'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
sheet1.merge_cells('A5:F5')

sheet1['A6'] = "Metric"
sheet1['B6'] = "Current"
sheet1['C6'] = "Target"
sheet1['D6'] = "Gap"
sheet1['E6'] = "Priority"
sheet1['F6'] = "Action"
for col in ['A', 'B', 'C', 'D', 'E', 'F']:
    sheet1[f'{col}6'].font = Font(bold=True)
    sheet1[f'{col}6'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

sheet1['A7'] = "Banks Cracked (Top 15)"
sheet1['B7'] = 6
sheet1['C7'] = 15
sheet1['D7'] = "=C7-B7"
sheet1['E7'] = "High"
sheet1['F7'] = "Implement next 9 banks"

sheet1['A8'] = "Market Coverage (Account %)"
sheet1['B8'] = "50%"
sheet1['C8'] = "80%"
sheet1['D8'] = "=C8-B8"
sheet1['E8'] = "High"
sheet1['F8'] = "Reach 80% coverage"

sheet1['A9'] = "Average Accuracy"
sheet1['B9'] = "92%"
sheet1['C9'] = "95%"
sheet1['D9'] = "=C9-B9"
sheet1['E9'] = "Medium"
sheet1['F9'] = "Improve to 95%"

sheet1['A10'] = "Sample PDFs Tested"
sheet1['B10'] = 0
sheet1['C10'] = 29
sheet1['D10'] = "=C10-B10"
sheet1['E10'] = "High"
sheet1['F10'] = "Test all 29 PDFs"

# Immediate Action Items
sheet1['A12'] = "IMMEDIATE ACTION ITEMS"
sheet1['A12'].font = Font(bold=True, size=14)
sheet1['A12'].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
sheet1.merge_cells('A12:F12')

sheet1['A13'] = "Priority"
sheet1['B13'] = "Action Item"
sheet1['C13'] = "Owner"
sheet1['D13'] = "Due Date"
sheet1['E13'] = "Status"
sheet1['F13'] = "Impact"
for col in ['A', 'B', 'C', 'D', 'E', 'F']:
    sheet1[f'{col}13'].font = Font(bold=True)
    sheet1[f'{col}13'].fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")

sheet1['A14'] = "P1"
sheet1['B14'] = "Test 29 sample PDFs to identify banks"
sheet1['C14'] = "Tech Team"
sheet1['D14'] = (datetime.now() + timedelta(days=3)).strftime("%Y-%m-%d")
sheet1['E14'] = "Not Started"
sheet1['F14'] = "High"

sheet1['A15'] = "P1"
sheet1['B15'] = "Complete ICICI accuracy testing"
sheet1['C15'] = "Dev Team"
sheet1['D15'] = (datetime.now() + timedelta(days=7)).strftime("%Y-%m-%d")
sheet1['E15'] = "In Progress"
sheet1['F15'] = "High"

sheet1['A16'] = "P1"
sheet1['B16'] = "Complete Axis accuracy testing"
sheet1['C16'] = "Dev Team"
sheet1['D16'] = (datetime.now() + timedelta(days=7)).strftime("%Y-%m-%d")
sheet1['E16'] = "In Progress"
sheet1['F16'] = "High"

sheet1['A17'] = "P2"
sheet1['B17'] = "Implement PNB (PSU priority)"
sheet1['C17'] = "Dev Team"
sheet1['D17'] = (datetime.now() + timedelta(days=30)).strftime("%Y-%m-%d")
sheet1['E17'] = "Not Started"
sheet1['F17'] = "High"

sheet1['A18'] = "P2"
sheet1['B18'] = "Implement Bank of Baroda"
sheet1['C18'] = "Dev Team"
sheet1['D18'] = (datetime.now() + timedelta(days=45)).strftime("%Y-%m-%d")
sheet1['E18'] = "Not Started"
sheet1['F18'] = "High"

# Column widths
sheet1.column_dimensions['A'].width = 12
sheet1.column_dimensions['B'].width = 35
sheet1.column_dimensions['C'].width = 15
sheet1.column_dimensions['D'].width = 12
sheet1.column_dimensions['E'].width = 12
sheet1.column_dimensions['F'].width = 12

# Sheet 2: Bank Coverage by Category
sheet2 = wb.create_sheet("Bank Coverage by Category")

sheet2['A1'] = "INDIAN BANKS - COMPLETE LIST"
sheet2['A1'].font = Font(bold=True, size=16, color="FFFFFF")
sheet2['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
sheet2['A1'].alignment = Alignment(horizontal="center")
sheet2.merge_cells('A1:G1')

# Headers
sheet2['A2'] = "Category"
sheet2['B2'] = "Bank Name"
sheet2['C2'] = "Market Share"
sheet2['D2'] = "Airco Status"
sheet2['E2'] = "Priority Score"
sheet2['F2'] = "Sample PDFs"
sheet2['G2'] = "Decision"
for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
    sheet2[f'{col}2'].font = Font(bold=True)
    sheet2[f'{col}2'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

# Public Sector Banks
public_sector = [
    ("SBI", "Very High", "Implemented", 9.5, "Yes"),
    ("PNB", "High", "Not Started", 8.5, "No"),
    ("Bank of Baroda", "High", "Not Started", 8.5, "No"),
    ("Canara Bank", "Medium", "Not Started", 7.0, "No"),
    ("Union Bank of India", "Medium", "Not Started", 7.0, "No"),
    ("Bank of Maharashtra", "Low", "Not Started", 5.0, "No"),
    ("Indian Bank", "Low", "Not Started", 5.0, "No"),
    ("Indian Overseas Bank", "Low", "Not Started", 4.0, "No"),
    ("Central Bank of India", "Low", "Not Started", 4.0, "No"),
    ("UCO Bank", "Low", "Not Started", 4.0, "No"),
    ("Bank of India", "Low", "Not Started", 4.0, "No"),
    ("Punjab & Sind Bank", "Very Low", "Not Started", 3.0, "No")
]

for i, (bank, market_share, status, priority, has_sample) in enumerate(public_sector, start=3):
    sheet2[f'A{i}'] = "Public Sector"
    sheet2[f'B{i}'] = bank
    sheet2[f'C{i}'] = market_share
    sheet2[f'D{i}'] = status
    sheet2[f'E{i}'] = priority
    sheet2[f'F{i}'] = has_sample
    sheet2[f'G{i}'] = "=IF(AND(E{i}>=8, F{i}='Yes'), 'Implement Next', IF(E{i}>=8, 'High Priority', IF(E{i}>=5, 'Medium Priority', 'Low Priority')))"

# Private Sector Banks
private_sector = [
    ("HDFC Bank", "Very High", "Complete", 10, "Yes"),
    ("ICICI Bank", "Very High", "Implemented", 9.5, "Yes"),
    ("Axis Bank", "Very High", "Implemented", 9.0, "Yes"),
    ("Kotak Mahindra Bank", "Very High", "Implemented", 9.0, "Yes"),
    ("IndusInd Bank", "High", "Not Started", 8.5, "No"),
    ("Yes Bank", "Medium", "Not Started", 7.0, "No"),
    ("IDFC FIRST Bank", "High", "Not Started", 8.0, "No"),
    ("Federal Bank", "Medium", "Not Started", 6.5, "No"),
    ("South Indian Bank", "Medium", "Not Started", 6.0, "No"),
    ("Karur Vysya Bank", "Low", "Not Started", 5.5, "No"),
    ("RBL Bank", "Medium", "Not Started", 6.5, "No"),
    ("Dhanlaxmi Bank", "Low", "Not Started", 5.0, "No"),
    ("City Union Bank", "Low", "Not Started", 5.0, "No"),
    ("Karnataka Bank", "Low", "Not Started", 5.0, "No"),
    ("Lakshmi Vilas Bank", "Very Low", "Not Started", 3.0, "No")
]

start_idx = 3 + len(public_sector) + 1
for i, (bank, market_share, status, priority, has_sample) in enumerate(private_sector, start=start_idx):
    sheet2[f'A{i}'] = "Private Sector"
    sheet2[f'B{i}'] = bank
    sheet2[f'C{i}'] = market_share
    sheet2[f'D{i}'] = status
    sheet2[f'E{i}'] = priority
    sheet2[f'F{i}'] = has_sample
    sheet2[f'G{i}'] = "=IF(AND(E{i}>=8, F{i}='Yes'), 'Implement Next', IF(E{i}>=8, 'High Priority', IF(E{i}>=5, 'Medium Priority', 'Low Priority')))"

# Foreign Banks
foreign_banks = [
    ("HSBC", "High", "Listed", 7.5, "No"),
    ("Citibank", "High", "Not Started", 7.5, "No"),
    ("Standard Chartered", "High", "Not Started", 7.5, "No"),
    ("DBS Bank", "Medium", "Not Started", 6.5, "No"),
    ("Deutsche Bank", "Medium", "Not Started", 6.0, "No"),
    ("Barclays", "Low", "Not Started", 5.5, "No"),
    ("ABN AMRO", "Low", "Not Started", 5.0, "No")
]

start_idx = start_idx + len(private_sector) + 1
for i, (bank, market_share, status, priority, has_sample) in enumerate(foreign_banks, start=start_idx):
    sheet2[f'A{i}'] = "Foreign"
    sheet2[f'B{i}'] = bank
    sheet2[f'C{i}'] = market_share
    sheet2[f'D{i}'] = status
    sheet2[f'E{i}'] = priority
    sheet2[f'F{i}'] = has_sample
    sheet2[f'G{i}'] = "=IF(AND(E{i}>=8, F{i}='Yes'), 'Implement Next', IF(E{i}>=8, 'High Priority', IF(E{i}>=5, 'Medium Priority', 'Low Priority')))"

# Column widths
sheet2.column_dimensions['A'].width = 18
sheet2.column_dimensions['B'].width = 28
sheet2.column_dimensions['C'].width = 14
sheet2.column_dimensions['D'].width = 15
sheet2.column_dimensions['E'].width = 14
sheet2.column_dimensions['F'].width = 14
sheet2.column_dimensions['G'].width = 18

# Sheet 3: Implementation Status & Accuracy
sheet3 = wb.create_sheet("Implementation Status")

sheet3['A1'] = "AIRCO BANK IMPLEMENTATION STATUS"
sheet3['A1'].font = Font(bold=True, size=16, color="FFFFFF")
sheet3['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
sheet3['A1'].alignment = Alignment(horizontal="center")
sheet3.merge_cells('A1:H1')

# Headers
sheet3['A2'] = "Category"
sheet3['B2'] = "Bank Name"
sheet3['C2'] = "Status"
sheet3['D2'] = "Accuracy"
sheet3['E2'] = "Modules (10)"
sheet3['F2'] = "Next Action"
sheet3['G2'] = "Owner"
sheet3['H2'] = "Target Date"
for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
    sheet3[f'{col}2'].font = Font(bold=True)
    sheet3[f'{col}2'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

# Implemented Banks
implemented_data = [
    ("Private Sector", "HDFC Bank", "Complete", "100%", 10, "Production ready", "Dev Team", "Done"),
    ("Private Sector", "ICICI Bank", "Implemented", "88.9%", 8, "Accuracy testing", "Dev Team", (datetime.now() + timedelta(days=7)).strftime("%Y-%m-%d")),
    ("Private Sector", "Axis Bank", "Implemented", "88.9%", 8, "Accuracy testing", "Dev Team", (datetime.now() + timedelta(days=7)).strftime("%Y-%m-%d")),
    ("Private Sector", "Kotak Bank", "Implemented", "TBD", 6, "Add AI fallback", "Dev Team", (datetime.now() + timedelta(days=14)).strftime("%Y-%m-%d")),
    ("Public Sector", "SBI", "Implemented", "TBD", 6, "Add AI fallback", "Dev Team", (datetime.now() + timedelta(days=14)).strftime("%Y-%m-%d")),
    ("Foreign", "HSBC", "Listed", "TBD", 0, "Start implementation", "Dev Team", (datetime.now() + timedelta(days=60)).strftime("%Y-%m-%d"))
]

for i, (category, bank, status, accuracy, modules, next_action, owner, target) in enumerate(implemented_data, start=3):
    sheet3[f'A{i}'] = category
    sheet3[f'B{i}'] = bank
    sheet3[f'C{i}'] = status
    sheet3[f'D{i}'] = accuracy
    sheet3[f'E{i}'] = modules
    sheet3[f'F{i}'] = next_action
    sheet3[f'G{i}'] = owner
    sheet3[f'H{i}'] = target

# Column widths
sheet3.column_dimensions['A'].width = 18
sheet3.column_dimensions['B'].width = 22
sheet3.column_dimensions['C'].width = 15
sheet3.column_dimensions['D'].width = 12
sheet3.column_dimensions['E'].width = 14
sheet3.column_dimensions['F'].width = 20
sheet3.column_dimensions['G'].width = 14
sheet3.column_dimensions['H'].width = 12

# Sheet 4: Priority Roadmap
sheet4 = wb.create_sheet("Priority Roadmap")

sheet4['A1'] = "IMPLEMENTATION ROADMAP - PHASES"
sheet4['A1'].font = Font(bold=True, size=16, color="FFFFFF")
sheet4['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
sheet4['A1'].alignment = Alignment(horizontal="center")
sheet4.merge_cells('A1:I1')

# Headers
sheet4['A2'] = "Phase"
sheet4['B2'] = "Bank Name"
sheet4['C2'] = "Category"
sheet4['D2'] = "Start Date"
sheet4['E2'] = "End Date"
sheet4['F2'] = "Effort (Days)"
sheet4['G2'] = "Team"
sheet4['H2'] = "Status"
sheet4['I2'] = "Milestone"
for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
    sheet4[f'{col}2'].font = Font(bold=True)
    sheet4[f'{col}2'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

# Phase 1: Current Banks (Complete)
phase1_data = [
    ("P1 - Current", "HDFC Bank", "Private Sector", "Done", "Done", 21, "Dev Team", "Complete", "Production"),
    ("P1 - Current", "ICICI Bank", "Private Sector", "Done", (datetime.now() + timedelta(days=7)).strftime("%Y-%m-%d"), 14, "Dev Team", "In Progress", "Testing"),
    ("P1 - Current", "Axis Bank", "Private Sector", "Done", (datetime.now() + timedelta(days=7)).strftime("%Y-%m-%d"), 14, "Dev Team", "In Progress", "Testing"),
    ("P1 - Current", "Kotak Bank", "Private Sector", "Done", (datetime.now() + timedelta(days=14)).strftime("%Y-%m-%d"), 10, "Dev Team", "In Progress", "AI Fallback"),
    ("P1 - Current", "SBI", "Public Sector", "Done", (datetime.now() + timedelta(days=14)).strftime("%Y-%m-%d"), 10, "Dev Team", "In Progress", "AI Fallback")
]

for i, (phase, bank, category, start, end, effort, team, status, milestone) in enumerate(phase1_data, start=3):
    sheet4[f'A{i}'] = phase
    sheet4[f'B{i}'] = bank
    sheet4[f'C{i}'] = category
    sheet4[f'D{i}'] = start
    sheet4[f'E{i}'] = end
    sheet4[f'F{i}'] = effort
    sheet4[f'G{i}'] = team
    sheet4[f'H{i}'] = status
    sheet4[f'I{i}'] = milestone

# Phase 2: High Priority
start_idx = 3 + len(phase1_data) + 1
phase2_data = [
    ("P2 - High", "PNB", "Public Sector", (datetime.now() + timedelta(days=15)).strftime("%Y-%m-%d"), (datetime.now() + timedelta(days=30)).strftime("%Y-%m-%d"), 14, "Dev Team", "Not Started", "Parser"),
    ("P2 - High", "Bank of Baroda", "Public Sector", (datetime.now() + timedelta(days=30)).strftime("%Y-%m-%d"), (datetime.now() + timedelta(days=45)).strftime("%Y-%m-%d"), 14, "Dev Team", "Not Started", "Parser"),
    ("P2 - High", "IndusInd Bank", "Private Sector", (datetime.now() + timedelta(days=45)).strftime("%Y-%m-%d"), (datetime.now() + timedelta(days=55)).strftime("%Y-%m-%d"), 10, "Dev Team", "Not Started", "Parser"),
    ("P2 - High", "IDFC FIRST Bank", "Private Sector", (datetime.now() + timedelta(days=55)).strftime("%Y-%m-%d"), (datetime.now() + timedelta(days=65)).strftime("%Y-%m-%d"), 10, "Dev Team", "Not Started", "Parser")
]

for i, (phase, bank, category, start, end, effort, team, status, milestone) in enumerate(phase2_data, start=start_idx):
    sheet4[f'A{i}'] = phase
    sheet4[f'B{i}'] = bank
    sheet4[f'C{i}'] = category
    sheet4[f'D{i}'] = start
    sheet4[f'E{i}'] = end
    sheet4[f'F{i}'] = effort
    sheet4[f'G{i}'] = team
    sheet4[f'H{i}'] = status
    sheet4[f'I{i}'] = milestone

# Column widths
sheet4.column_dimensions['A'].width = 15
sheet4.column_dimensions['B'].width = 22
sheet4.column_dimensions['C'].width = 18
sheet4.column_dimensions['D'].width = 12
sheet4.column_dimensions['E'].width = 12
sheet4.column_dimensions['F'].width = 14
sheet4.column_dimensions['G'].width = 12
sheet4.column_dimensions['H'].width = 12
sheet4.column_dimensions['I'].width = 15

# Sheet 5: Sample PDF Testing
sheet5 = wb.create_sheet("Sample PDF Testing")

sheet5['A1'] = "SAMPLE PDF TESTING TRACKER (29 PDFs)"
sheet5['A1'].font = Font(bold=True, size=16, color="FFFFFF")
sheet5['A1'].fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
sheet5['A1'].alignment = Alignment(horizontal="center")
sheet5.merge_cells('A1:G1')

sheet5['A2'] = "PDF Filename"
sheet5['B2'] = "Identified Bank"
sheet5['C2'] = "Test Status"
sheet5['D2'] = "Accuracy Result"
sheet5['E2'] = "Issues Found"
sheet5['F2'] = "Test Date"
sheet5['G2'] = "Tester"
for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
    sheet5[f'{col}2'].font = Font(bold=True)
    sheet5[f'{col}2'].fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")

# Sample PDF data (placeholder for 29 PDFs)
sample_pdfs = [
    ("50100154418680_1686406381909.pdf", "Unknown", "Not Tested", "N/A", "N/A", "", ""),
    ("Mypasiccstatement_1681198287513.pdf", "Unknown", "Not Tested", "N/A", "N/A", "", ""),
    ("AcctStatement_XX8762_24052023(1).pdf", "Unknown", "Not Tested", "N/A", "N/A", "", ""),
    ("HdfcStatement_XX9887_09082023(1).pdf", "HDFC", "Not Tested", "N/A", "N/A", "", ""),
    ("IDFCFIRSTBankstatement_10072076528(1).pdf", "IDFC FIRST", "Not Tested", "N/A", "N/A", "", "")
]

for i, (pdf, bank, status, accuracy, issues, date, tester) in enumerate(sample_pdfs, start=3):
    sheet5[f'A{i}'] = pdf
    sheet5[f'B{i}'] = bank
    sheet5[f'C{i}'] = status
    sheet5[f'D{i}'] = accuracy
    sheet5[f'E{i}'] = issues
    sheet5[f'F{i}'] = date
    sheet5[f'G{i}'] = tester

# Summary row
summary_idx = 3 + len(sample_pdfs) + 1
sheet5[f'A{summary_idx}'] = "TOTAL"
sheet5[f'A{summary_idx}'].font = Font(bold=True)
sheet5[f'C{summary_idx}'] = "=COUNTIF(C3:C{summary_idx-1}, 'Tested')"
sheet5[f'C{summary_idx}'].font = Font(bold=True)
sheet5[f'D{summary_idx}'] = "=COUNTA(D3:D{summary_idx-1})"
sheet5[f'D{summary_idx}'].font = Font(bold=True)

sheet5.column_dimensions['A'].width = 40
sheet5.column_dimensions['B'].width = 20
sheet5.column_dimensions['C'].width = 15
sheet5.column_dimensions['D'].width = 15
sheet5.column_dimensions['E'].width = 25
sheet5.column_dimensions['F'].width = 12
sheet5.column_dimensions['G'].width = 15

# Sheet 6: Resource Planning
sheet6 = wb.create_sheet("Resource Planning")

sheet6['A1'] = "RESOURCE ALLOCATION PLANNING"
sheet6['A1'].font = Font(bold=True, size=16, color="FFFFFF")
sheet6['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
sheet6['A1'].alignment = Alignment(horizontal="center")
sheet6.merge_cells('A1:H1')

sheet6['A2'] = "Month"
sheet6['B2'] = "Banks Planned"
sheet6['C2'] = "Dev Resources"
sheet6['D2'] = "QA Resources"
sheet6['E2'] = "Total Effort (Days)"
sheet6['F2'] = "Deliverable"
sheet6['G2'] = "Budget (INR)"
sheet6['H2'] = "Risk"
for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
    sheet6[f'{col}2'].font = Font(bold=True)
    sheet6[f'{col}2'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

resource_data = [
    ("May 2026", "ICICI, Axis accuracy", 2, 1, 21, "Testing complete", 50000, "Low"),
    ("June 2026", "PNB", 2, 1, 14, "Parser complete", 40000, "Medium"),
    ("July 2026", "Bank of Baroda", 2, 1, 14, "Parser complete", 40000, "Medium"),
    ("August 2026", "IndusInd, IDFC", 2, 1, 20, "Parser complete", 55000, "Low"),
    ("September 2026", "Kotak, SBI AI", 2, 1, 20, "AI fallback complete", 55000, "Medium"),
    ("October 2026", "Standard Chartered", 2, 1, 12, "Parser complete", 35000, "High")
]

for i, (month, banks, dev, qa, effort, deliverable, budget, risk) in enumerate(resource_data, start=3):
    sheet6[f'A{i}'] = month
    sheet6[f'B{i}'] = banks
    sheet6[f'C{i}'] = dev
    sheet6[f'D{i}'] = qa
    sheet6[f'E{i}'] = effort
    sheet6[f'F{i}'] = deliverable
    sheet6[f'G{i}'] = budget
    sheet6[f'H{i}'] = risk

sheet6.column_dimensions['A'].width = 12
sheet6.column_dimensions['B'].width = 25
sheet6.column_dimensions['C'].width = 14
sheet6.column_dimensions['D'].width = 14
sheet6.column_dimensions['E'].width = 18
sheet6.column_dimensions['F'].width = 20
sheet6.column_dimensions['G'].width = 12
sheet6.column_dimensions['H'].width = 10

# Sheet 7: Decision Matrix
sheet7 = wb.create_sheet("Decision Matrix")

sheet7['A1'] = "BANK IMPLEMENTATION DECISION MATRIX"
sheet7['A1'].font = Font(bold=True, size=16, color="FFFFFF")
sheet7['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
sheet7['A1'].alignment = Alignment(horizontal="center")
sheet7.merge_cells('A1:J1')

sheet7['A2'] = "Bank Name"
sheet7['B2'] = "Market Share"
sheet7['C2'] = "Sample PDFs"
sheet7['D2'] = "Complexity"
sheet7['E2'] = "Priority Score"
sheet7['F2'] = "Est. Days"
sheet7['G2'] = "ROI"
sheet7['H2'] = "Risk Level"
sheet7['I2'] = "Decision"
sheet7['J2'] = "Timeline"
for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
    sheet7[f'{col}2'].font = Font(bold=True)
    sheet7[f'{col}2'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

decision_data = [
    ("PNB", "High", "No", "Medium", 8.5, 14, "High", "Medium", "Implement Q2", "30 days"),
    ("Bank of Baroda", "High", "No", "Medium", 8.5, 14, "High", "Medium", "Implement Q2", "45 days"),
    ("IndusInd Bank", "High", "No", "Low", 8.5, 10, "High", "Low", "Implement Q3", "10 days"),
    ("IDFC FIRST Bank", "High", "No", "Low", 8.0, 10, "High", "Low", "Implement Q3", "10 days"),
    ("Canara Bank", "Medium", "No", "Medium", 7.0, 14, "Medium", "Medium", "Evaluate", "60 days"),
    ("Union Bank", "Medium", "No", "Medium", 7.0, 14, "Medium", "Medium", "Evaluate", "60 days"),
    ("Standard Chartered", "High", "No", "High", 7.5, 12, "Medium", "High", "Evaluate Q4", "90 days"),
    ("Citibank", "High", "No", "High", 7.5, 12, "Medium", "High", "Evaluate Q4", "90 days")
]

for i, (bank, market, samples, complexity, priority, days, roi, risk, decision, timeline) in enumerate(decision_data, start=3):
    sheet7[f'A{i}'] = bank
    sheet7[f'B{i}'] = market
    sheet7[f'C{i}'] = samples
    sheet7[f'D{i}'] = complexity
    sheet7[f'E{i}'] = priority
    sheet7[f'F{i}'] = days
    sheet7[f'G{i}'] = roi
    sheet7[f'H{i}'] = risk
    sheet7[f'I{i}'] = decision
    sheet7[f'J{i}'] = timeline

sheet7.column_dimensions['A'].width = 22
sheet7.column_dimensions['B'].width = 14
sheet7.column_dimensions['C'].width = 14
sheet7.column_dimensions['D'].width = 14
sheet7.column_dimensions['E'].width = 14
sheet7.column_dimensions['F'].width = 12
sheet7.column_dimensions['G'].width = 10
sheet7.column_dimensions['H'].width = 12
sheet7.column_dimensions['I'].width = 15
sheet7.column_dimensions['J'].width = 12

# Save the workbook
output_path = "x:/FinTech SAAS/Airco Insights Fintech/Bank_Coverage_Tracker_v2.xlsx"
wb.save(output_path)
print(f"Excel tracker created successfully: {output_path}")
